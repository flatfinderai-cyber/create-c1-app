"""
FlatFinder Commercial — Toronto Full-Market Scraper (browser-use edition)
Platforms : Kijiji · Zumper · PadMapper · Craigslist · Rentals.ca · Apartments.ca
Coverage  : ALL unit types, ALL price ranges, ALL Toronto neighbourhoods
Output    : flatfinder_toronto.xlsx  (daily tab + All Listings tab)
            flatfinder_toronto_latest.csv  (Google Sheets import / API feed)
Run       : python flatfinder_scraper.py
Requires  : Set one of these env vars for the LLM backend:
              BROWSER_USE_API_KEY  — ChatBrowserUse (cheapest, $0.20/1M tokens)
              ANTHROPIC_API_KEY    — Claude claude-sonnet-4-6
              OPENAI_API_KEY       — GPT-4o mini
"""

import os, re, csv, logging, hashlib, asyncio, json
from datetime import date
from typing import Optional, List

from pydantic import BaseModel
from browser_use import Agent, Browser
from browser_use.browser.session import BrowserSession
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
CITY       = "Toronto"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE  = os.path.join(OUTPUT_DIR, "flatfinder_toronto.xlsx")
CSV_FILE   = os.path.join(OUTPUT_DIR, "flatfinder_toronto_latest.csv")

COLS = [
    "ID", "Source", "Title", "Price", "Bedrooms", "Bathrooms",
    "Type", "Neighbourhood", "Address", "Utilities", "Pets",
    "TTC_Access", "Available", "URL", "Description", "Date_Scraped"
]

COL_WIDTHS = {
    "ID": 10, "Source": 12, "Title": 46, "Price": 10,
    "Bedrooms": 10, "Bathrooms": 10, "Type": 16,
    "Neighbourhood": 22, "Address": 30, "Utilities": 12,
    "Pets": 8, "TTC_Access": 12, "Available": 14,
    "URL": 14, "Description": 44, "Date_Scraped": 14
}

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── STYLES ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="0D1B2A")
HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
EVEN_FILL = PatternFill("solid", start_color="F4F6FB")
ODD_FILL  = PatternFill("solid", start_color="FFFFFF")
UTIL_FILL = PatternFill("solid", start_color="D6F5E3")
PET_FILL  = PatternFill("solid", start_color="FFF3E0")
LINK_FONT = Font(name="Arial", color="1155CC", underline="single", size=9)
BOLD9     = Font(name="Arial", size=9, bold=True)
REG9      = Font(name="Arial", size=9)
GREY8     = Font(name="Arial", size=8, color="666666")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_W    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN      = Side(style="thin", color="D0D5E8")
BORDER    = Border(bottom=THIN)

# ── PYDANTIC MODELS FOR STRUCTURED EXTRACTION ─────────────────────────────────
class RawListing(BaseModel):
    title: str
    price: Optional[str] = None
    url: Optional[str] = None
    bedrooms: Optional[str] = None
    neighbourhood: Optional[str] = None
    utilities: Optional[str] = None
    pets: Optional[str] = None
    description: Optional[str] = None
    available: Optional[str] = None

class PlatformListings(BaseModel):
    listings: List[RawListing]

# ── PLATFORM TASKS ────────────────────────────────────────────────────────────
PLATFORMS = [
    {
        "source": "Kijiji",
        "task": (
            "Go to https://www.kijiji.ca/b-apartments-condos/city-of-toronto/c37l1700273 "
            "and extract ALL apartment rental listings shown on the page. "
            "For each listing collect: title, monthly price (digits only), full listing URL, "
            "bedrooms (e.g. Bachelor, 1-Bed, 2-Bed, 3-Bed), neighbourhood/location, "
            "utilities included (Yes/Partial/Check), pets allowed (Yes/No/?), "
            "and a brief description under 200 characters. "
            "Return all listings as a JSON object with a 'listings' array."
        ),
    },
    {
        "source": "Zumper",
        "task": (
            "Go to https://www.zumper.com/apartments-for-rent/toronto-on "
            "and extract all apartment rental listings. Scroll down to load more if needed. "
            "For each listing collect: title or address, monthly price (digits only), full URL, "
            "bedrooms, neighbourhood, utilities info, pets policy, brief description. "
            "Return all listings as a JSON object with a 'listings' array."
        ),
    },
    {
        "source": "PadMapper",
        "task": (
            "Go to https://www.padmapper.com/apartments/toronto-on "
            "and extract all visible apartment rental listings. "
            "Click on listing cards if needed to get details. "
            "For each listing collect: title or address, monthly price (digits only), full URL, "
            "bedrooms, neighbourhood, utilities info, pets policy, brief description. "
            "Return all listings as a JSON object with a 'listings' array."
        ),
    },
    {
        "source": "Craigslist",
        "task": (
            "Go to https://toronto.craigslist.org/search/tor/apa "
            "and extract all apartment rental listings on the page. "
            "For each listing collect: title, monthly price (digits only), full URL, "
            "bedrooms (infer from title if possible), neighbourhood, "
            "utilities info, pets policy, brief description. "
            "Return all listings as a JSON object with a 'listings' array."
        ),
    },
    {
        "source": "Rentals.ca",
        "task": (
            "Go to https://rentals.ca/toronto "
            "and extract all rental listings shown on the page. "
            "For each listing collect: building name or title, monthly price (digits only), full URL, "
            "bedrooms, neighbourhood, utilities info, pets policy, brief description. "
            "Return all listings as a JSON object with a 'listings' array."
        ),
    },
    {
        "source": "Apartments.ca",
        "task": (
            "Go to https://www.apartments.ca/toronto/ "
            "and extract all apartment rental listings shown on the page. "
            "For each listing collect: building name or title, monthly price (digits only), full URL, "
            "bedrooms, neighbourhood, utilities info, pets policy, brief description. "
            "Return all listings as a JSON object with a 'listings' array."
        ),
    },
]

# ── LLM FACTORY ───────────────────────────────────────────────────────────────
def get_llm():
    if os.getenv("BROWSER_USE_API_KEY"):
        from browser_use import ChatBrowserUse
        return ChatBrowserUse()
    elif os.getenv("ANTHROPIC_API_KEY"):
        from browser_use.llm import ChatAnthropic
        return ChatAnthropic(
            model="claude-sonnet-4-6",
            api_key=os.getenv("ANTHROPIC_API_KEY"),
        )
    elif os.getenv("OPENAI_API_KEY"):
        from browser_use.llm import ChatOpenAI
        return ChatOpenAI(
            model="gpt-4o-mini",
            api_key=os.getenv("OPENAI_API_KEY"),
        )
    else:
        raise ValueError(
            "No LLM API key found. Set one of:\n"
            "  BROWSER_USE_API_KEY  (ChatBrowserUse — cheapest)\n"
            "  ANTHROPIC_API_KEY    (Claude claude-sonnet-4-6)\n"
            "  OPENAI_API_KEY       (GPT-4o mini)"
        )

# ── HELPERS ───────────────────────────────────────────────────────────────────
def clean(t):
    return " ".join(str(t).strip().split()) if t else ""

def parse_price(text):
    if not text:
        return None
    m = re.search(r"[\d,]+", str(text).replace(",", ""))
    if m:
        try:
            return int(m.group().replace(",", ""))
        except ValueError:
            return None
    return None

def make_id(source, title, price):
    raw = f"{source}{title}{price}".encode()
    return hashlib.md5(raw).hexdigest()[:8].upper()

def normalize_beds(text: str) -> str:
    if not text:
        return "Unknown"
    t = str(text).lower().strip()
    if any(w in t for w in ["bach", "studio", "0 bed", "0br"]):
        return "Bachelor/Studio"
    for n in ["6", "5", "4", "3", "2", "1"]:
        if t == n or f"{n} bed" in t or f"{n}br" in t or f"{n}-bed" in t:
            return f"{n}-Bed"
    return detect_beds(text)

def detect_beds(text):
    t = str(text).lower()
    if any(w in t for w in ["bachelor", "studio", "bach", "0 bed"]):
        return "Bachelor/Studio"
    for n in ["5", "6", "7"]:
        if f"{n} bed" in t or f"{n}bed" in t or f"{n}br" in t:
            return f"{n}-Bed"
    for n, w in [("4", "four"), ("3", "three"), ("2", "two"), ("1", "one")]:
        if f"{n} bed" in t or f"{n}bed" in t or f"{n}br" in t or f"{w} bed" in t:
            return f"{n}-Bed"
    return "Unknown"

def detect_baths(text):
    t = str(text).lower()
    for n in ["4", "3", "2"]:
        if f"{n} bath" in t or f"{n}bath" in t:
            return n
    return "1" if "bath" in t else "?"

def normalize_utilities(text: str) -> str:
    if not text:
        return "Check"
    t = str(text).lower()
    if any(w in t for w in ["yes", "incl", "all", "included", "utilities included"]):
        return "Yes"
    if any(w in t for w in ["partial", "some", "heat only", "water only"]):
        return "Partial"
    return "Check"

def detect_utilities(text):
    t = str(text).lower()
    if any(w in t for w in ["all incl", "all-incl", "utilities incl", "all inclusive",
                              "heat incl", "hydro incl", "water incl", "bills incl",
                              "everything incl", "utilities included"]):
        return "Yes"
    if any(w in t for w in ["heat only", "water only", "hydro extra", "+ hydro",
                              "+ utilities", "utilities not", "hydro not incl"]):
        return "Partial"
    return "Check"

def normalize_pets(text: str) -> str:
    if not text:
        return "?"
    t = str(text).lower()
    if any(w in t for w in ["yes", "allow", "friendly", "ok", "welcome", "permitted"]):
        return "Yes"
    if any(w in t for w in ["no", "not allowed", "prohibit", "free", "no pets"]):
        return "No"
    return "?"

def detect_pets(text):
    t = str(text).lower()
    if any(w in t for w in ["pet friendly", "pets allowed", "pets ok", "dogs ok",
                              "cats ok", "pets welcome"]):
        return "Yes"
    if any(w in t for w in ["no pets", "pet free", "no dogs", "no cats"]):
        return "No"
    return "?"

def detect_ttc(text, address=""):
    t = (str(text) + " " + str(address)).lower()
    subway_keywords = [
        "subway", "ttc", "bloor-yonge", "spadina stn", "union stn",
        "osgoode", "st. patrick", "queen stn", "king stn", "dundas stn",
        "college stn", "wellesley", "sherbourne", "castle frank", "broadview",
        "chester", "pape", "donlands", "greenwood", "coxwell", "woodbine",
        "main street", "victoria park", "warden", "kennedy", "scarborough",
        "york mills", "sheppard", "wilson", "yorkdale", "lawrence",
        "eglinton", "davisville", "st. clair", "summerhill", "rosedale",
        "bay", "museum", "queens park", "st. george", "dupont", "spadina",
        "bathurst", "ossington", "dufferin", "lansdowne", "dundas west",
        "runnymede", "jane", "old mill", "kipling", "islington", "royal york",
        "high park", "keele", "finch", "line 1", "line 2",
        "steps to subway", "walk to subway", "min to subway",
        "near subway", "close to subway",
    ]
    if any(kw in t for kw in subway_keywords):
        return "Subway"
    streetcar = ["streetcar", "504 ", "505 ", "506 ", "509 ", "510 ", "511 ",
                 "512 ", "queen st", "king st", "college st", "dundas st",
                 "bathurst st", "carlton"]
    if any(kw in t for kw in streetcar):
        return "Streetcar"
    if any(kw in t for kw in ["bus", "ttc bus", "transit"]):
        return "Bus"
    return "?"

def detect_available(text):
    t = str(text).lower()
    patterns = [
        r"available\s+([\w\s,]+\d{4})",
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[\s.,-]+\d{1,2}[\s,]+\d{4}",
        r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}",
        r"(immediately|now|asap|right away)",
        r"march 1|april 1|may 1",
    ]
    for p in patterns:
        m = re.search(p, t)
        if m:
            return clean(m.group())[:20]
    return ""

# ── PLATFORM SCRAPER ──────────────────────────────────────────────────────────
def parse_agent_result(result) -> List[dict]:
    """Parse whatever the agent returns into a list of raw listing dicts."""
    if result is None:
        return []

    # Already a PlatformListings Pydantic model
    if isinstance(result, PlatformListings):
        return [l.model_dump() for l in result.listings]

    # A dict with a listings key
    if isinstance(result, dict):
        items = result.get("listings", [])
        return items if isinstance(items, list) else []

    # A JSON string
    if isinstance(result, str):
        try:
            data = json.loads(result)
            if isinstance(data, dict):
                items = data.get("listings", [])
                return items if isinstance(items, list) else []
            if isinstance(data, list):
                return data
        except json.JSONDecodeError:
            pass

    return []


def normalize_raw(raw: dict, source: str) -> Optional[dict]:
    """Convert a raw listing dict from the agent into our standard schema."""
    title = clean(raw.get("title", ""))
    if not title:
        return None

    price_raw = raw.get("price") or ""
    price = parse_price(str(price_raw))
    url = clean(raw.get("url") or "")
    beds_raw = clean(raw.get("bedrooms") or "")
    desc_raw = clean(raw.get("description") or "")
    neighbourhood = clean(raw.get("neighbourhood") or CITY)
    utilities_raw = raw.get("utilities") or ""
    pets_raw = raw.get("pets") or ""
    available_raw = raw.get("available") or ""

    combined = title + " " + beds_raw + " " + desc_raw

    bedrooms = normalize_beds(beds_raw) if beds_raw else detect_beds(combined)
    utilities = normalize_utilities(utilities_raw) if utilities_raw else detect_utilities(combined)
    pets = normalize_pets(pets_raw) if pets_raw else detect_pets(combined)
    available = clean(available_raw)[:20] if available_raw else detect_available(combined)

    return {
        "ID": make_id(source, title, price),
        "Source": source,
        "Title": title,
        "Price": price,
        "Bedrooms": bedrooms,
        "Bathrooms": detect_baths(combined),
        "Type": "Apartment",
        "Neighbourhood": neighbourhood,
        "Address": title if any(c.isdigit() for c in title[:5]) else "",
        "Utilities": utilities,
        "Pets": pets,
        "TTC_Access": detect_ttc(combined, neighbourhood),
        "Available": available,
        "URL": url,
        "Description": desc_raw[:220],
        "Date_Scraped": str(date.today()),
    }


async def scrape_platform(platform: dict, llm) -> List[dict]:
    source = platform["source"]
    task = platform["task"]
    log.info(f"Starting {source}...")

    try:
        browser = BrowserSession(
            headless=True,
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            minimum_wait_page_load_time=1.5,
            wait_for_network_idle_page_load_time=3.0,
        )
        agent = Agent(
            task=task,
            llm=llm,
            browser=browser,
            output_model_schema=PlatformListings,
            max_failures=3,
            use_vision=True,
            enable_planning=True,
        )
        history = await agent.run(max_steps=30)
        result = history.final_result()

        raw_listings = parse_agent_result(result)
        listings = []
        for raw in raw_listings:
            normalized = normalize_raw(raw, source)
            if normalized:
                listings.append(normalized)

        log.info(f"{source}: {len(listings)} listings extracted")
        return listings

    except Exception as e:
        log.error(f"{source} failed: {e}")
        return []

# ── DEDUP ─────────────────────────────────────────────────────────────────────
def deduplicate(listings):
    seen, out = set(), []
    for l in listings:
        key = l["ID"]
        if key not in seen:
            seen.add(key)
            out.append(l)
    return out

# ── XLSX WRITER ───────────────────────────────────────────────────────────────
def style_row(ws, ri, l, fill):
    ws.row_dimensions[ri].height = 18
    for ci, col in enumerate(COLS, 1):
        val  = l.get(col, "")
        cell = ws.cell(ri, ci)
        cell.border = BORDER

        if col == "Price":
            cell.value         = l.get("Price")
            cell.number_format = '"$"#,##0'
            cell.font          = Font(name="Arial", size=9, bold=True)
            cell.alignment     = CENTER
            cell.fill          = fill

        elif col == "URL" and val:
            cell.value     = "Open"
            cell.hyperlink = val
            cell.font      = LINK_FONT
            cell.alignment = CENTER
            cell.fill      = fill

        elif col == "Utilities":
            cell.value = val
            cell.font  = Font(name="Arial", size=9, bold=(val == "Yes"),
                              color="1A7A3C" if val == "Yes" else
                                    "E65100" if val == "Partial" else "555555")
            cell.fill      = UTIL_FILL if val == "Yes" else fill
            cell.alignment = CENTER

        elif col == "Pets":
            cell.value = val
            cell.font  = Font(name="Arial", size=9,
                              color="1A7A3C" if val == "Yes" else
                                    "C62828" if val == "No" else "555555")
            cell.fill      = PET_FILL if val == "Yes" else fill
            cell.alignment = CENTER

        elif col == "TTC_Access":
            color = {"Subway": "1155CC", "Streetcar": "6A1B9A", "Bus": "2E7D32"}.get(val, "555555")
            cell.value = val
            cell.font  = Font(name="Arial", size=9, color=color,
                              bold=(val in ("Subway", "Streetcar")))
            cell.fill      = fill
            cell.alignment = CENTER

        elif col == "Title":
            cell.value     = val
            cell.font      = BOLD9
            cell.fill      = fill
            cell.alignment = LEFT_W

        elif col == "Description":
            cell.value     = val
            cell.font      = GREY8
            cell.fill      = fill
            cell.alignment = LEFT_W

        elif col in ("ID", "Source", "Bedrooms", "Bathrooms", "Type", "Available", "Date_Scraped"):
            cell.value     = val
            cell.font      = REG9
            cell.fill      = fill
            cell.alignment = CENTER

        else:
            cell.value     = val
            cell.font      = REG9
            cell.fill      = fill
            cell.alignment = LEFT_W


def write_sheet(ws, listings, sheet_label):
    ws.row_dimensions[1].height = 22
    for ci, col in enumerate(COLS, 1):
        cell = ws.cell(1, ci, col)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[col]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    for ri, l in enumerate(listings, 2):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        style_row(ws, ri, l, fill)

    last = len(listings) + 2
    ws.cell(last, 1, f"Total: {len(listings)}").font = BOLD9
    ws.cell(last, 3, f"=COUNTA(C2:C{last-1})").font  = REG9
    ws.cell(last, 4, "Avg:").font                    = REG9
    ws.cell(last, 5, f"=IFERROR(AVERAGE(D2:D{last-1}),\"-\")").number_format = '"$"#,##0'
    ws.cell(last, 5).font = REG9


def write_xlsx(listings):
    today_str = str(date.today())

    if os.path.exists(XLSX_FILE):
        wb = load_workbook(XLSX_FILE)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if today_str in wb.sheetnames:
        del wb[today_str]
    ws_today = wb.create_sheet(title=today_str, index=0)
    write_sheet(ws_today, listings, today_str)

    ALL = "All Listings"
    if ALL not in wb.sheetnames:
        wa = wb.create_sheet(ALL)
        write_sheet(wa, [], ALL)
    else:
        wa = wb[ALL]

    existing_ids = set()
    for row in wa.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            existing_ids.add(row[0])

    new_rows = [l for l in listings if l["ID"] not in existing_ids]
    start_ri = wa.max_row + 1
    for ri, l in enumerate(new_rows, start_ri):
        fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        wa.row_dimensions[ri].height = 18
        style_row(wa, ri, l, fill)

    STATS = "Stats"
    if STATS in wb.sheetnames:
        del wb[STATS]
    ws_stats = wb.create_sheet(STATS)
    write_stats_sheet(ws_stats, listings, today_str)

    wb.save(XLSX_FILE)
    log.info(f"XLSX saved → {XLSX_FILE}  |  today: {len(listings)}  |  new to All: {len(new_rows)}")


def write_stats_sheet(ws, listings, today_str):
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    title_cell = ws.cell(1, 1, f"FlatFinder Toronto — {today_str}")
    title_cell.font = Font(name="Arial", size=13, bold=True, color="0D1B2A")
    ws.merge_cells("A1:B1")

    rows = [
        ("Total listings today",   len(listings)),
        ("Sources scraped",         len(set(l["Source"] for l in listings))),
        ("Bachelor/Studio",         sum(1 for l in listings if "Bach" in str(l.get("Bedrooms", "")) or "Studio" in str(l.get("Bedrooms", "")))),
        ("1-Bedroom",               sum(1 for l in listings if l.get("Bedrooms") == "1-Bed")),
        ("2-Bedroom",               sum(1 for l in listings if l.get("Bedrooms") == "2-Bed")),
        ("3-Bedroom+",              sum(1 for l in listings if l.get("Bedrooms") in ("3-Bed", "4-Bed", "5-Bed", "6-Bed"))),
        ("Utilities Included",      sum(1 for l in listings if l.get("Utilities") == "Yes")),
        ("Pet Friendly",            sum(1 for l in listings if l.get("Pets") == "Yes")),
        ("Subway Access",           sum(1 for l in listings if l.get("TTC_Access") == "Subway")),
        ("Avg Price (all)",         int(sum(l["Price"] for l in listings if l.get("Price")) / max(1, sum(1 for l in listings if l.get("Price"))))),
        ("Min Price",               min((l["Price"] for l in listings if l.get("Price")), default=0)),
        ("Max Price",               max((l["Price"] for l in listings if l.get("Price")), default=0)),
    ]

    for ri, (label, value) in enumerate(rows, 3):
        lc = ws.cell(ri, 1, label)
        vc = ws.cell(ri, 2, value)
        lc.font = REG9
        vc.font = BOLD9
        lc.fill = vc.fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        lc.alignment = vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.number_format = '"$"#,##0' if "Price" in label else "General"
        lc.border = vc.border = BORDER

    ws.cell(len(rows) + 5, 1, "By Source").font = BOLD9
    by_source = {}
    for l in listings:
        by_source[l["Source"]] = by_source.get(l["Source"], 0) + 1
    for ri, (src, cnt) in enumerate(sorted(by_source.items(), key=lambda x: -x[1]), len(rows) + 6):
        ws.cell(ri, 1, src).font = REG9
        ws.cell(ri, 2, cnt).font = BOLD9
        ws.cell(ri, 1).fill = ws.cell(ri, 2).fill = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        ws.cell(ri, 1).border = ws.cell(ri, 2).border = BORDER


# ── CSV WRITER ────────────────────────────────────────────────────────────────
def write_csv(listings):
    with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLS)
        writer.writeheader()
        for l in listings:
            writer.writerow({col: l.get(col, "") for col in COLS})
    log.info(f"CSV  saved → {CSV_FILE}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
async def async_main():
    log.info("=" * 55)
    log.info("  FlatFinder — browser-use Scraper — START")
    log.info("=" * 55)

    llm = get_llm()

    all_listings = []
    for platform in PLATFORMS:
        listings = await scrape_platform(platform, llm)
        all_listings.extend(listings)

    all_listings = deduplicate(all_listings)
    all_listings.sort(key=lambda x: (x.get("Price") or 999999))

    log.info(f"Total unique: {len(all_listings)}")
    write_xlsx(all_listings)
    write_csv(all_listings)
    log.info("  FlatFinder — browser-use Scraper — DONE")
    log.info("=" * 55)


def main():
    asyncio.run(async_main())


if __name__ == "__main__":
    main()
